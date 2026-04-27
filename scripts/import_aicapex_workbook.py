#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import math
import os
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


YEARS = list(range(2026, 2046))
START_YEAR = 2026
MID_YEAR = 2030
END_YEAR = 2045
PROJECT_ROOT = Path(__file__).resolve().parents[1]


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        raise FileNotFoundError(f"Missing env file: {path}")
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        env[key] = value
    return env


def mysql_client() -> str:
    configured = os.environ.get("MYSQL_CLI")
    if configured:
        return configured
    found = shutil.which("mysql")
    if not found:
        raise RuntimeError("mysql CLI not found on PATH")
    return found


def mysql_args(env: dict[str, str], database: str | None = None) -> list[str]:
    args = [
        mysql_client(),
        "--host",
        env["MYSQL_HOST"],
        "--port",
        str(env.get("MYSQL_PORT", "3306")),
        "--user",
        env["MYSQL_USER"],
        "--default-character-set=utf8mb4",
        "--batch",
        "--raw",
    ]
    if database:
        args.append(database)
    return args


def mysql_execute(sql: str, env: dict[str, str], database: str | None = None) -> str:
    proc_env = os.environ.copy()
    proc_env["MYSQL_PWD"] = env["MYSQL_PASSWORD"]
    proc = subprocess.run(
        mysql_args(env, database),
        input=sql,
        text=True,
        capture_output=True,
        env=proc_env,
    )
    if proc.returncode != 0:
        stderr = proc.stderr.replace(env["MYSQL_PASSWORD"], "***")
        raise RuntimeError(f"MySQL command failed:\n{stderr}")
    return proc.stdout


def sql_ident(name: str) -> str:
    return "`" + name.replace("`", "``") + "`"


def sql_value(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return "NULL"
        return repr(value)
    text = str(value).replace("\x00", "")
    text = text.replace("\\", "\\\\").replace("'", "''")
    return "'" + text + "'"


def insert_rows(
    env: dict[str, str],
    database: str,
    table: str,
    columns: list[str],
    rows: list[tuple[Any, ...]],
    batch_size: int = 700,
) -> None:
    if not rows:
        return
    col_sql = ", ".join(sql_ident(c) for c in columns)
    for offset in range(0, len(rows), batch_size):
        chunk = rows[offset : offset + batch_size]
        values_sql = ",\n".join(
            "(" + ", ".join(sql_value(v) for v in row) + ")" for row in chunk
        )
        sql = f"INSERT INTO {sql_ident(table)} ({col_sql}) VALUES\n{values_sql};"
        mysql_execute(sql, env, database)


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def interpolate(v2026: float, v2030: float, v2045: float, year: int) -> float:
    if year <= MID_YEAR:
        return v2026 + (v2030 - v2026) * (year - START_YEAR) / (MID_YEAR - START_YEAR)
    return v2030 + (v2045 - v2030) * (year - MID_YEAR) / (END_YEAR - MID_YEAR)


def as_float(value: Any, default: float | None = None) -> float | None:
    if value is None:
        return default
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return default


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").upper().replace("=", "").replace("()", "").strip()
    return text == "TRUE"


def workbook_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def read_simple_anchors(ws, start_row: int = 4) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for r in range(start_row, ws.max_row + 1):
        item = ws.cell(r, 1).value
        if not item:
            continue
        a = ws.cell(r, 23).value
        b = ws.cell(r, 24).value
        c = ws.cell(r, 25).value
        if not all(is_number(x) for x in (a, b, c)):
            continue
        rows.append(
            {
                "item": str(item),
                "target_2026": float(a),
                "target_2030": float(b),
                "target_2045": float(c),
                "rationale": ws.cell(r, 26).value,
                "source_refs": ws.cell(r, 27).value,
            }
        )
    return rows


def read_grouped_anchors(ws, start_row: int = 4) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    owner: str | None = None
    for r in range(start_row, ws.max_row + 1):
        item = ws.cell(r, 1).value
        if not item:
            continue
        a = ws.cell(r, 23).value
        b = ws.cell(r, 24).value
        c = ws.cell(r, 25).value
        if all(is_number(x) for x in (a, b, c)):
            if owner is None:
                continue
            rows.append(
                {
                    "owner": owner,
                    "item": str(item),
                    "target_2026": float(a),
                    "target_2030": float(b),
                    "target_2045": float(c),
                    "rationale": ws.cell(r, 26).value,
                    "source_refs": ws.cell(r, 27).value,
                }
            )
        else:
            owner = str(item)
    return rows


def values_by_year_from_anchors(rows: list[dict[str, Any]]) -> dict[str, dict[int, float]]:
    out: dict[str, dict[int, float]] = {}
    for row in rows:
        out[row["item"]] = {
            year: interpolate(row["target_2026"], row["target_2030"], row["target_2045"], year)
            for year in YEARS
        }
    return out


def grouped_values_by_year(rows: list[dict[str, Any]]) -> dict[str, dict[str, dict[int, float]]]:
    out: dict[str, dict[str, dict[int, float]]] = {}
    for row in rows:
        out.setdefault(row["owner"], {})[row["item"]] = {
            year: interpolate(row["target_2026"], row["target_2030"], row["target_2045"], year)
            for year in YEARS
        }
    return out


def create_schema(env: dict[str, str]) -> None:
    database = env["MYSQL_DATABASE"]
    mysql_execute(
        f"CREATE DATABASE IF NOT EXISTS {sql_ident(database)} "
        "DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci;",
        env,
    )
    ddl = """
CREATE TABLE IF NOT EXISTS model_runs (
  run_id VARCHAR(96) PRIMARY KEY,
  workbook_file VARCHAR(255) NOT NULL,
  workbook_sha256 CHAR(64) NOT NULL,
  forecast_start_year INT NOT NULL,
  forecast_end_year INT NOT NULL,
  imported_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS source_register (
  run_id VARCHAR(96) NOT NULL,
  source_id VARCHAR(32) NOT NULL,
  theme VARCHAR(128),
  key_fact TEXT,
  url TEXT,
  source_type VARCHAR(64),
  notes TEXT,
  PRIMARY KEY (run_id, source_id),
  CONSTRAINT fk_sources_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS driver_contributions (
  run_id VARCHAR(96) NOT NULL,
  year_num INT NOT NULL,
  driver_name VARCHAR(160) NOT NULL,
  contribution DECIMAL(20,10) NOT NULL,
  PRIMARY KEY (run_id, year_num, driver_name),
  CONSTRAINT fk_drivers_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS global_forecast (
  run_id VARCHAR(96) NOT NULL,
  year_num INT NOT NULL,
  total_usd_bn DOUBLE NOT NULL,
  yoy_growth DOUBLE,
  PRIMARY KEY (run_id, year_num),
  CONSTRAINT fk_global_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS share_assumptions (
  run_id VARCHAR(96) NOT NULL,
  dimension_name VARCHAR(32) NOT NULL,
  item_name VARCHAR(191) NOT NULL,
  target_2026 DOUBLE,
  target_2030 DOUBLE,
  target_2045 DOUBLE,
  rationale TEXT,
  source_refs TEXT,
  PRIMARY KEY (run_id, dimension_name, item_name),
  CONSTRAINT fk_share_assumptions_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS forecast_totals (
  run_id VARCHAR(96) NOT NULL,
  dimension_name VARCHAR(32) NOT NULL,
  item_name VARCHAR(191) NOT NULL,
  year_num INT NOT NULL,
  share_of_global DOUBLE,
  amount_usd_bn DOUBLE NOT NULL,
  PRIMARY KEY (run_id, dimension_name, item_name, year_num),
  CONSTRAINT fk_forecast_totals_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS memory_storage_split_assumptions (
  run_id VARCHAR(96) NOT NULL,
  split_family VARCHAR(191) NOT NULL,
  target_2026 DOUBLE,
  target_2030 DOUBLE,
  target_2045 DOUBLE,
  applied_to VARCHAR(128),
  source_logic TEXT,
  PRIMARY KEY (run_id, split_family),
  CONSTRAINT fk_mem_storage_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS component_useful_life (
  run_id VARCHAR(96) NOT NULL,
  component_name VARCHAR(191) NOT NULL,
  useful_life_years DOUBLE,
  annual_depreciation_proxy DOUBLE,
  note TEXT,
  source_refs TEXT,
  PRIMARY KEY (run_id, component_name),
  CONSTRAINT fk_life_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS component_tilt_assumptions (
  run_id VARCHAR(96) NOT NULL,
  entity_type VARCHAR(32) NOT NULL,
  entity_name VARCHAR(128) NOT NULL,
  component_name VARCHAR(191) NOT NULL,
  target_2026 DOUBLE,
  target_2030 DOUBLE,
  target_2045 DOUBLE,
  rationale TEXT,
  source_refs TEXT,
  PRIMARY KEY (run_id, entity_type, entity_name, component_name),
  CONSTRAINT fk_component_tilt_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS entity_component_forecast (
  run_id VARCHAR(96) NOT NULL,
  entity_type VARCHAR(32) NOT NULL,
  entity_name VARCHAR(128) NOT NULL,
  component_name VARCHAR(191) NOT NULL,
  year_num INT NOT NULL,
  tilt_multiplier DOUBLE,
  component_share DOUBLE,
  amount_usd_bn DOUBLE,
  PRIMARY KEY (run_id, entity_type, entity_name, component_name, year_num),
  CONSTRAINT fk_entity_component_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS company_country_shares (
  run_id VARCHAR(96) NOT NULL,
  company_name VARCHAR(128) NOT NULL,
  country_name VARCHAR(128) NOT NULL,
  year_num INT NOT NULL,
  tilt_multiplier DOUBLE,
  prior_share DOUBLE,
  PRIMARY KEY (run_id, company_name, country_name, year_num),
  CONSTRAINT fk_company_country_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS ras_factors (
  run_id VARCHAR(96) NOT NULL,
  factor_type VARCHAR(32) NOT NULL,
  item_name VARCHAR(128) NOT NULL,
  year_num INT NOT NULL,
  factor_value DOUBLE NOT NULL,
  PRIMARY KEY (run_id, factor_type, item_name, year_num),
  CONSTRAINT fk_ras_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS country_company_allocation (
  run_id VARCHAR(96) NOT NULL,
  country_name VARCHAR(128) NOT NULL,
  company_name VARCHAR(128) NOT NULL,
  year_num INT NOT NULL,
  prior_share DOUBLE,
  country_factor DOUBLE,
  company_factor DOUBLE,
  reconciled_share DOUBLE,
  allocated_total_usd_bn DOUBLE,
  PRIMARY KEY (run_id, country_name, company_name, year_num),
  CONSTRAINT fk_country_company_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS country_company_component_forecast (
  run_id VARCHAR(96) NOT NULL,
  country_name VARCHAR(128) NOT NULL,
  company_name VARCHAR(128) NOT NULL,
  component_name VARCHAR(191) NOT NULL,
  year_num INT NOT NULL,
  amount_usd_bn DOUBLE,
  PRIMARY KEY (run_id, country_name, company_name, component_name, year_num),
  CONSTRAINT fk_ccc_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS funding_source_metadata (
  run_id VARCHAR(96) NOT NULL,
  funding_source VARCHAR(191) NOT NULL,
  economic_cost BOOLEAN,
  debt_like BOOLEAN,
  cash_interest BOOLEAN,
  principal_roll_down BOOLEAN,
  note TEXT,
  PRIMARY KEY (run_id, funding_source),
  CONSTRAINT fk_funding_meta_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS company_funding_assumptions (
  run_id VARCHAR(96) NOT NULL,
  company_name VARCHAR(128) NOT NULL,
  funding_source VARCHAR(191) NOT NULL,
  metric_name VARCHAR(32) NOT NULL,
  target_2026 DOUBLE,
  target_2030 DOUBLE,
  target_2045 DOUBLE,
  rationale TEXT,
  source_refs TEXT,
  PRIMARY KEY (run_id, company_name, funding_source, metric_name),
  CONSTRAINT fk_funding_assumptions_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS company_funding_terms (
  run_id VARCHAR(96) NOT NULL,
  company_name VARCHAR(128) NOT NULL,
  funding_source VARCHAR(191) NOT NULL,
  year_num INT NOT NULL,
  share_of_annual_investment DOUBLE,
  amount_funded_usd_bn DOUBLE,
  cost_rate DOUBLE,
  tenor_years DOUBLE,
  economic_cost BOOLEAN,
  debt_like BOOLEAN,
  cash_interest BOOLEAN,
  economic_cost_current_flow_usd_bn DOUBLE,
  cash_interest_current_flow_usd_bn DOUBLE,
  PRIMARY KEY (run_id, company_name, funding_source, year_num),
  CONSTRAINT fk_funding_terms_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS company_roic_spread (
  run_id VARCHAR(96) NOT NULL,
  company_name VARCHAR(128) NOT NULL,
  year_num INT NOT NULL,
  spread DOUBLE,
  PRIMARY KEY (run_id, company_name, year_num),
  CONSTRAINT fk_roic_spread_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;

CREATE TABLE IF NOT EXISTS company_finance_roic (
  run_id VARCHAR(96) NOT NULL,
  company_name VARCHAR(128) NOT NULL,
  year_num INT NOT NULL,
  annual_investment_usd_bn DOUBLE,
  weighted_all_in_funding_cost DOUBLE,
  weighted_debt_rate DOUBLE,
  weighted_debt_tenor_years DOUBLE,
  debt_funded_share DOUBLE,
  beginning_invested_capital_usd_bn DOUBLE,
  depreciation_rate DOUBLE,
  depreciation_expense_usd_bn DOUBLE,
  ending_invested_capital_usd_bn DOUBLE,
  average_invested_capital_usd_bn DOUBLE,
  implied_roic DOUBLE,
  implied_nopat_usd_bn DOUBLE,
  economic_funding_cost_usd_bn DOUBLE,
  roic_spread DOUBLE,
  eva_usd_bn DOUBLE,
  beginning_debt_usd_bn DOUBLE,
  new_debt_funded_investment_usd_bn DOUBLE,
  debt_amortization_usd_bn DOUBLE,
  ending_debt_usd_bn DOUBLE,
  average_debt_usd_bn DOUBLE,
  cash_interest_burden_usd_bn DOUBLE,
  interest_coverage DOUBLE,
  weighted_avg_useful_life_years DOUBLE,
  PRIMARY KEY (run_id, company_name, year_num),
  CONSTRAINT fk_finance_roic_run FOREIGN KEY (run_id) REFERENCES model_runs(run_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
"""
    # The remote import user may not have REFERENCES privilege. The model is
    # versioned by run_id and the importer deletes/reloads one run at a time, so
    # foreign keys are useful but not required for this database.
    ddl = "\n".join(line for line in ddl.splitlines() if "CONSTRAINT fk_" not in line)
    ddl = ddl.replace(",\n) ENGINE", "\n) ENGINE")
    mysql_execute(ddl, env, database)


def import_workbook(env: dict[str, str]) -> dict[str, int]:
    database = env["MYSQL_DATABASE"]
    excel_path = Path(env.get("EXCEL_PATH", "")).expanduser()
    if not excel_path.is_absolute():
        excel_path = PROJECT_ROOT / excel_path
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    sha = workbook_hash(excel_path)
    run_id = f"{excel_path.stem[:48]}_{sha[:10]}"

    wb = load_workbook(excel_path, data_only=False, read_only=True)

    mysql_execute(f"DELETE FROM model_runs WHERE run_id = {sql_value(run_id)};", env, database)
    insert_rows(
        env,
        database,
        "model_runs",
        ["run_id", "workbook_file", "workbook_sha256", "forecast_start_year", "forecast_end_year"],
        [(run_id, excel_path.name, sha, START_YEAR, END_YEAR)],
    )

    counts: dict[str, int] = {}

    # Sources.
    rows = []
    ws = wb["Sources"]
    for r in range(4, ws.max_row + 1):
        source_id = ws.cell(r, 1).value
        if not source_id:
            continue
        rows.append(
            (
                run_id,
                source_id,
                ws.cell(r, 2).value,
                ws.cell(r, 3).value,
                ws.cell(r, 4).value,
                ws.cell(r, 5).value,
                ws.cell(r, 6).value,
            )
        )
    insert_rows(env, database, "source_register", ["run_id", "source_id", "theme", "key_fact", "url", "source_type", "notes"], rows)
    counts["source_register"] = len(rows)

    # Drivers and global forecast.
    ws = wb["Drivers"]
    driver_rows = []
    growth: dict[int, float] = {}
    for col, year in enumerate(range(2027, 2046), start=2):
        total = 0.0
        for r in range(4, 9):
            driver = str(ws.cell(r, 1).value)
            value = as_float(ws.cell(r, col).value, 0.0) or 0.0
            driver_rows.append((run_id, year, driver, value))
            total += value
        growth[year] = total
    insert_rows(env, database, "driver_contributions", ["run_id", "year_num", "driver_name", "contribution"], driver_rows)
    counts["driver_contributions"] = len(driver_rows)

    global_amount: dict[int, float] = {2026: 1450.0}
    for year in range(2027, 2046):
        global_amount[year] = global_amount[year - 1] * (1.0 + growth[year])
    global_rows = []
    for year in YEARS:
        yoy = None if year == 2026 else global_amount[year] / global_amount[year - 1] - 1.0
        global_rows.append((run_id, year, global_amount[year], yoy))
    insert_rows(env, database, "global_forecast", ["run_id", "year_num", "total_usd_bn", "yoy_growth"], global_rows)
    counts["global_forecast"] = len(global_rows)

    country_assumptions = read_simple_anchors(wb["Country_Shares"])
    company_assumptions = read_simple_anchors(wb["Company_Shares"])
    component_assumptions = read_simple_anchors(wb["Component_Shares"])
    share_rows = []
    for dimension, assumptions in [
        ("country", country_assumptions),
        ("company", company_assumptions),
        ("component", component_assumptions),
    ]:
        for row in assumptions:
            share_rows.append(
                (
                    run_id,
                    dimension,
                    row["item"],
                    row["target_2026"],
                    row["target_2030"],
                    row["target_2045"],
                    row["rationale"],
                    row["source_refs"],
                )
            )
    insert_rows(
        env,
        database,
        "share_assumptions",
        ["run_id", "dimension_name", "item_name", "target_2026", "target_2030", "target_2045", "rationale", "source_refs"],
        share_rows,
    )
    counts["share_assumptions"] = len(share_rows)

    country_shares = values_by_year_from_anchors(country_assumptions)
    company_shares = values_by_year_from_anchors(company_assumptions)
    component_shares = values_by_year_from_anchors(component_assumptions)
    countries = [r["item"] for r in country_assumptions]
    companies = [r["item"] for r in company_assumptions]
    components = [r["item"] for r in component_assumptions]

    forecast_rows = []
    forecast_rows.extend((run_id, "global", "Global total", year, 1.0, global_amount[year]) for year in YEARS)
    country_amounts: dict[str, dict[int, float]] = {}
    company_amounts: dict[str, dict[int, float]] = {}
    component_amounts: dict[str, dict[int, float]] = {}
    for country in countries:
        country_amounts[country] = {}
        for year in YEARS:
            amount = country_shares[country][year] * global_amount[year]
            country_amounts[country][year] = amount
            forecast_rows.append((run_id, "country", country, year, country_shares[country][year], amount))
    for company in companies:
        company_amounts[company] = {}
        for year in YEARS:
            amount = company_shares[company][year] * global_amount[year]
            company_amounts[company][year] = amount
            forecast_rows.append((run_id, "company", company, year, company_shares[company][year], amount))
    for component in components:
        component_amounts[component] = {}
        for year in YEARS:
            amount = component_shares[component][year] * global_amount[year]
            component_amounts[component][year] = amount
            forecast_rows.append((run_id, "component", component, year, component_shares[component][year], amount))
    insert_rows(env, database, "forecast_totals", ["run_id", "dimension_name", "item_name", "year_num", "share_of_global", "amount_usd_bn"], forecast_rows)
    counts["forecast_totals"] = len(forecast_rows)

    # Memory/storage split assumptions.
    ws = wb["Mem_Storage_Split_Assumptions"]
    rows = []
    for r in range(4, ws.max_row + 1):
        family = ws.cell(r, 1).value
        if not family:
            continue
        rows.append((run_id, family, ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 6).value))
    insert_rows(env, database, "memory_storage_split_assumptions", ["run_id", "split_family", "target_2026", "target_2030", "target_2045", "applied_to", "source_logic"], rows)
    counts["memory_storage_split_assumptions"] = len(rows)

    # Useful life.
    ws = wb["Component_Useful_Life"]
    life: dict[str, float] = {}
    dep_proxy: dict[str, float] = {}
    rows = []
    for r in range(5, ws.max_row + 1):
        component = ws.cell(r, 1).value
        useful_life = as_float(ws.cell(r, 2).value)
        if not component or useful_life is None:
            continue
        dep = 1.0 / useful_life if useful_life else None
        life[str(component)] = useful_life
        dep_proxy[str(component)] = dep or 0.0
        rows.append((run_id, component, useful_life, dep, ws.cell(r, 4).value, ws.cell(r, 5).value))
    insert_rows(env, database, "component_useful_life", ["run_id", "component_name", "useful_life_years", "annual_depreciation_proxy", "note", "source_refs"], rows)
    counts["component_useful_life"] = len(rows)

    # Country/company component tilts and normalized shares.
    country_tilt_rows = read_grouped_anchors(wb["Country_Component_Tilt"])
    company_tilt_rows = read_grouped_anchors(wb["Company_Component_Tilt"])
    rows = []
    for entity_type, tilt_rows in [("country", country_tilt_rows), ("company", company_tilt_rows)]:
        for row in tilt_rows:
            rows.append((run_id, entity_type, row["owner"], row["item"], row["target_2026"], row["target_2030"], row["target_2045"], row["rationale"], row["source_refs"]))
    insert_rows(env, database, "component_tilt_assumptions", ["run_id", "entity_type", "entity_name", "component_name", "target_2026", "target_2030", "target_2045", "rationale", "source_refs"], rows)
    counts["component_tilt_assumptions"] = len(rows)

    country_tilts = grouped_values_by_year(country_tilt_rows)
    company_tilts = grouped_values_by_year(company_tilt_rows)

    def normalized_component_shares(
        owners: list[str],
        totals: dict[str, dict[int, float]],
        tilts: dict[str, dict[str, dict[int, float]]],
        entity_type: str,
    ) -> tuple[dict[str, dict[str, dict[int, float]]], list[tuple[Any, ...]]]:
        shares: dict[str, dict[str, dict[int, float]]] = {}
        rows_out: list[tuple[Any, ...]] = []
        for owner in owners:
            shares[owner] = {}
            for year in YEARS:
                denom = sum(component_shares[c][year] * tilts[owner][c][year] for c in components)
                for c in components:
                    tilt = tilts[owner][c][year]
                    share = (component_shares[c][year] * tilt / denom) if denom else 0.0
                    shares[owner].setdefault(c, {})[year] = share
                    rows_out.append((run_id, entity_type, owner, c, year, tilt, share, totals[owner][year] * share))
        return shares, rows_out

    country_component_shares, country_component_rows = normalized_component_shares(countries, country_amounts, country_tilts, "country")
    company_component_shares, company_component_rows = normalized_component_shares(companies, company_amounts, company_tilts, "company")
    insert_rows(env, database, "entity_component_forecast", ["run_id", "entity_type", "entity_name", "component_name", "year_num", "tilt_multiplier", "component_share", "amount_usd_bn"], country_component_rows + company_component_rows)
    counts["entity_component_forecast"] = len(country_component_rows) + len(company_component_rows)

    # Company-country shares.
    company_country_tilt_rows = read_grouped_anchors(wb["Company_Country_Tilt"])
    company_country_tilts = grouped_values_by_year(company_country_tilt_rows)
    company_country_shares: dict[str, dict[str, dict[int, float]]] = {}
    rows = []
    for company in companies:
        company_country_shares[company] = {}
        for year in YEARS:
            denom = sum(country_shares[country][year] * company_country_tilts[company][country][year] for country in countries)
            for country in countries:
                tilt = company_country_tilts[company][country][year]
                share = (country_shares[country][year] * tilt / denom) if denom else 0.0
                company_country_shares[company].setdefault(country, {})[year] = share
                rows.append((run_id, company, country, year, tilt, share))
    insert_rows(env, database, "company_country_shares", ["run_id", "company_name", "country_name", "year_num", "tilt_multiplier", "prior_share"], rows)
    counts["company_country_shares"] = len(rows)

    # RAS factors.
    ws = wb["Country_Company_RAS_Factors"]
    country_factors: dict[str, dict[int, float]] = {}
    company_factors: dict[str, dict[int, float]] = {}
    rows = []
    for r in range(6, 6 + len(countries)):
        country = str(ws.cell(r, 1).value)
        country_factors[country] = {}
        for col, year in enumerate(YEARS, start=2):
            val = as_float(ws.cell(r, col).value, 1.0) or 1.0
            country_factors[country][year] = val
            rows.append((run_id, "country", country, year, val))
    company_start = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").startswith("Company scaling"):
            company_start = r + 2
            break
    if company_start is None:
        raise RuntimeError("Could not locate company RAS factors")
    for r in range(company_start, company_start + len(companies)):
        company = str(ws.cell(r, 1).value)
        company_factors[company] = {}
        for col, year in enumerate(YEARS, start=2):
            val = as_float(ws.cell(r, col).value, 1.0) or 1.0
            company_factors[company][year] = val
            rows.append((run_id, "company", company, year, val))
    insert_rows(env, database, "ras_factors", ["run_id", "factor_type", "item_name", "year_num", "factor_value"], rows)
    counts["ras_factors"] = len(rows)

    bridge_ws = wb["Country_Co_Comp_Bridge"]
    company_weight = as_float(bridge_ws["AE5"].value, 0.6) or 0.6
    country_weight = as_float(bridge_ws["AE6"].value, 0.4) or 0.4
    allocation_rows = []
    ccc_rows = []
    for year in YEARS:
        for country in countries:
            for company in companies:
                prior = company_country_shares[company][country][year]
                cf = country_factors[country][year]
                kf = company_factors[company][year]
                reconciled = prior * cf * kf
                allocated = company_amounts[company][year] * reconciled
                allocation_rows.append((run_id, country, company, year, prior, cf, kf, reconciled, allocated))
                for component in components:
                    share = (
                        company_weight * company_component_shares[company][component][year]
                        + country_weight * country_component_shares[country][component][year]
                    )
                    ccc_rows.append((run_id, country, company, component, year, allocated * share))
    insert_rows(env, database, "country_company_allocation", ["run_id", "country_name", "company_name", "year_num", "prior_share", "country_factor", "company_factor", "reconciled_share", "allocated_total_usd_bn"], allocation_rows)
    insert_rows(env, database, "country_company_component_forecast", ["run_id", "country_name", "company_name", "component_name", "year_num", "amount_usd_bn"], ccc_rows)
    counts["country_company_allocation"] = len(allocation_rows)
    counts["country_company_component_forecast"] = len(ccc_rows)

    # Funding metadata and terms.
    ws = wb["Funding_Source_Metadata"]
    metadata: dict[str, dict[str, Any]] = {}
    rows = []
    for r in range(5, ws.max_row + 1):
        source = ws.cell(r, 1).value
        if not source:
            continue
        source = str(source)
        item = {
            "economic_cost": as_bool(ws.cell(r, 2).value),
            "debt_like": as_bool(ws.cell(r, 3).value),
            "cash_interest": as_bool(ws.cell(r, 4).value),
            "principal_roll_down": as_bool(ws.cell(r, 5).value),
            "note": ws.cell(r, 6).value,
        }
        metadata[source] = item
        rows.append((run_id, source, item["economic_cost"], item["debt_like"], item["cash_interest"], item["principal_roll_down"], item["note"]))
    insert_rows(env, database, "funding_source_metadata", ["run_id", "funding_source", "economic_cost", "debt_like", "cash_interest", "principal_roll_down", "note"], rows)
    counts["funding_source_metadata"] = len(rows)

    funding_share_rows = read_grouped_anchors(wb["Company_Funding_Shares"])
    funding_cost_rows = read_grouped_anchors(wb["Company_Funding_Cost_Rates"])
    funding_tenor_rows = read_grouped_anchors(wb["Company_Funding_Tenors"])
    funding_shares = grouped_values_by_year(funding_share_rows)
    funding_costs = grouped_values_by_year(funding_cost_rows)
    funding_tenors = grouped_values_by_year(funding_tenor_rows)
    funding_sources = list(metadata.keys())

    rows = []
    for metric, metric_rows in [("share", funding_share_rows), ("cost_rate", funding_cost_rows), ("tenor_years", funding_tenor_rows)]:
        for row in metric_rows:
            rows.append((run_id, row["owner"], row["item"], metric, row["target_2026"], row["target_2030"], row["target_2045"], row["rationale"], row["source_refs"]))
    insert_rows(env, database, "company_funding_assumptions", ["run_id", "company_name", "funding_source", "metric_name", "target_2026", "target_2030", "target_2045", "rationale", "source_refs"], rows)
    counts["company_funding_assumptions"] = len(rows)

    funding_term_rows = []
    for company in companies:
        for source in funding_sources:
            for year in YEARS:
                share = funding_shares[company][source][year]
                amount = company_amounts[company][year] * share
                cost = funding_costs[company][source][year]
                tenor = funding_tenors[company][source][year]
                meta = metadata[source]
                economic_cost_flow = amount * cost if meta["economic_cost"] else 0.0
                cash_interest_flow = amount * cost if meta["cash_interest"] else 0.0
                funding_term_rows.append((run_id, company, source, year, share, amount, cost, tenor, meta["economic_cost"], meta["debt_like"], meta["cash_interest"], economic_cost_flow, cash_interest_flow))
    insert_rows(env, database, "company_funding_terms", ["run_id", "company_name", "funding_source", "year_num", "share_of_annual_investment", "amount_funded_usd_bn", "cost_rate", "tenor_years", "economic_cost", "debt_like", "cash_interest", "economic_cost_current_flow_usd_bn", "cash_interest_current_flow_usd_bn"], funding_term_rows)
    counts["company_funding_terms"] = len(funding_term_rows)

    roic_spread_rows = read_simple_anchors(wb["Company_ROIC_Spread"])
    roic_spread_values = values_by_year_from_anchors(roic_spread_rows)
    rows = []
    for company in companies:
        for year in YEARS:
            rows.append((run_id, company, year, roic_spread_values[company][year]))
    insert_rows(env, database, "company_roic_spread", ["run_id", "company_name", "year_num", "spread"], rows)
    counts["company_roic_spread"] = len(rows)

    finance_rows = []
    for company in companies:
        beginning_invested = 0.0
        beginning_debt = 0.0
        for year in YEARS:
            annual_investment = company_amounts[company][year]
            weighted_all_in = sum(funding_shares[company][s][year] * funding_costs[company][s][year] for s in funding_sources)
            debt_sources = [s for s in funding_sources if metadata[s]["debt_like"]]
            debt_share = sum(funding_shares[company][s][year] for s in debt_sources)
            if debt_share:
                weighted_debt_rate = sum(funding_shares[company][s][year] * funding_costs[company][s][year] for s in debt_sources) / debt_share
                weighted_debt_tenor = sum(funding_shares[company][s][year] * funding_tenors[company][s][year] for s in debt_sources) / debt_share
            else:
                weighted_debt_rate = 0.0
                weighted_debt_tenor = 0.0
            depreciation_rate = sum(company_component_shares[company][c][year] * dep_proxy[c] for c in components)
            depreciation_expense = beginning_invested * depreciation_rate
            ending_invested = beginning_invested - depreciation_expense + annual_investment
            average_invested = (beginning_invested + ending_invested) / 2.0
            spread = roic_spread_values[company][year]
            implied_roic = weighted_all_in + spread
            implied_nopat = average_invested * implied_roic
            economic_funding_cost = average_invested * weighted_all_in
            eva = average_invested * spread
            new_debt = annual_investment * debt_share
            debt_amortization = beginning_debt / weighted_debt_tenor if weighted_debt_tenor else 0.0
            ending_debt = beginning_debt - debt_amortization + new_debt
            average_debt = (beginning_debt + ending_debt) / 2.0
            interest_burden = average_debt * weighted_debt_rate
            interest_coverage = implied_nopat / interest_burden if interest_burden else None
            weighted_life = sum(company_component_shares[company][c][year] * life[c] for c in components)
            finance_rows.append(
                (
                    run_id,
                    company,
                    year,
                    annual_investment,
                    weighted_all_in,
                    weighted_debt_rate,
                    weighted_debt_tenor,
                    debt_share,
                    beginning_invested,
                    depreciation_rate,
                    depreciation_expense,
                    ending_invested,
                    average_invested,
                    implied_roic,
                    implied_nopat,
                    economic_funding_cost,
                    spread,
                    eva,
                    beginning_debt,
                    new_debt,
                    debt_amortization,
                    ending_debt,
                    average_debt,
                    interest_burden,
                    interest_coverage,
                    weighted_life,
                )
            )
            beginning_invested = ending_invested
            beginning_debt = ending_debt
    insert_rows(
        env,
        database,
        "company_finance_roic",
        [
            "run_id",
            "company_name",
            "year_num",
            "annual_investment_usd_bn",
            "weighted_all_in_funding_cost",
            "weighted_debt_rate",
            "weighted_debt_tenor_years",
            "debt_funded_share",
            "beginning_invested_capital_usd_bn",
            "depreciation_rate",
            "depreciation_expense_usd_bn",
            "ending_invested_capital_usd_bn",
            "average_invested_capital_usd_bn",
            "implied_roic",
            "implied_nopat_usd_bn",
            "economic_funding_cost_usd_bn",
            "roic_spread",
            "eva_usd_bn",
            "beginning_debt_usd_bn",
            "new_debt_funded_investment_usd_bn",
            "debt_amortization_usd_bn",
            "ending_debt_usd_bn",
            "average_debt_usd_bn",
            "cash_interest_burden_usd_bn",
            "interest_coverage",
            "weighted_avg_useful_life_years",
        ],
        finance_rows,
    )
    counts["company_finance_roic"] = len(finance_rows)

    counts["run_id"] = run_id  # type: ignore[assignment]
    return counts


def main() -> int:
    env = load_env(PROJECT_ROOT / ".env")
    required = ["MYSQL_HOST", "MYSQL_PORT", "MYSQL_USER", "MYSQL_PASSWORD", "MYSQL_DATABASE"]
    missing = [key for key in required if not env.get(key)]
    if missing:
        raise RuntimeError(f"Missing required env keys: {', '.join(missing)}")

    create_schema(env)
    counts = import_workbook(env)
    run_id = counts.pop("run_id")
    print(f"Imported model run: {run_id}")
    for table, count in sorted(counts.items()):
        print(f"{table}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
