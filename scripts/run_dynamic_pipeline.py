#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import math
import os
import shutil
import subprocess
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
YEARS = list(range(2027, 2046))
DRIVER_ROWS = {
    "AI adoption / demand diffusion": 4,
    "Liquidity / financing": 5,
    "Energy & geopolitics": 6,
    "Trade / regionalization": 7,
    "Cycle / digestion": 8,
}
FRED_BASE_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
SEC_COMPANY_FACTS_URL = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
SEC_CAPEX_CONCEPTS = [
    "PaymentsToAcquirePropertyPlantAndEquipment",
    "PaymentsToAcquireProductiveAssets",
    "PropertyPlantAndEquipmentAdditions",
]


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def utc_stamp(now: datetime | None = None) -> str:
    return (now or utc_now()).strftime("%Y%m%dT%H%M%SZ")


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def merged_env() -> dict[str, str]:
    env = load_env(PROJECT_ROOT / ".env")
    for key in [
        "EXCEL_PATH",
        "EXTERNAL_SOURCE_CONFIG_PATH",
        "MODEL_ARCHIVE_DIR",
        "PIPELINE_TIMEOUT_SECONDS",
        "PIPELINE_FETCH_CONCURRENCY",
        "SEC_USER_AGENT",
        "PYTHON_BIN",
    ]:
        if os.environ.get(key):
            env[key] = str(os.environ[key])
    return env


def project_path(value: str | Path) -> Path:
    path = Path(value).expanduser()
    return path if path.is_absolute() else PROJECT_ROOT / path


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(float(value)) or math.isinf(float(value)):
            return None
        return float(value)
    text = str(value).strip()
    if not text or text == ".":
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def parse_iso_date(value: Any) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        return None


def shift_year(value: date, years: int) -> date:
    try:
        return value.replace(year=value.year + years)
    except ValueError:
        return value.replace(year=value.year + years, month=2, day=28)


def days_between(left: date, right: date) -> int:
    return abs((left - right).days)


def load_config(env: dict[str, str]) -> dict[str, Any]:
    path = project_path(env.get("EXTERNAL_SOURCE_CONFIG_PATH", "config/external_data_sources.json"))
    if not path.exists():
        raise FileNotFoundError(f"Missing external source config: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def curl_fetch_text(url: str, timeout_seconds: int, headers: dict[str, str] | None = None) -> str:
    curl = shutil.which("curl")
    if not curl:
        raise FileNotFoundError("curl not found")
    args = [curl, "--http1.1", "--fail", "--location", "--silent", "--show-error", "--max-time", str(timeout_seconds)]
    for key, value in (headers or {}).items():
        if key.lower() == "user-agent":
            args.extend(["--user-agent", value])
        else:
            args.extend(["--header", f"{key}: {value}"])
    args.append(url)
    proc = subprocess.run(args, text=True, capture_output=True)
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or f"curl exited with {proc.returncode}")
    return proc.stdout


def urllib_fetch_text(url: str, timeout_seconds: int, headers: dict[str, str] | None = None) -> tuple[str, dict[str, str]]:
    req = urllib.request.Request(url, headers=headers or {})
    with urllib.request.urlopen(req, timeout=timeout_seconds) as response:
        body = response.read().decode("utf-8", errors="replace")
        response_headers = {k.lower(): v for k, v in response.headers.items()}
    return body, response_headers


def fetch_text(url: str, timeout_seconds: int, headers: dict[str, str] | None = None) -> tuple[str, dict[str, str]]:
    last_error: Exception | None = None
    for attempt in range(1, 3):
        try:
            if shutil.which("curl"):
                return curl_fetch_text(url, timeout_seconds, headers), {}
            return urllib_fetch_text(url, timeout_seconds, headers)
        except (TimeoutError, urllib.error.URLError, RuntimeError, FileNotFoundError) as error:
            last_error = error
            if attempt < 2:
                time.sleep(attempt)
    assert last_error is not None
    raise last_error


def write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value, encoding="utf-8")


def write_json(path: Path, value: Any) -> None:
    write_text(path, f"{json.dumps(value, indent=2, sort_keys=True)}\n")


def nearest_prior_value(rows: list[tuple[date, float]], target: date) -> tuple[date, float] | None:
    candidates = [row for row in rows if row[0] <= target]
    if not candidates:
        return None
    return max(candidates, key=lambda row: row[0])


def parse_fred_csv(text: str, series_id: str) -> dict[str, Any]:
    reader = csv.DictReader(text.splitlines())
    rows: list[tuple[date, float]] = []
    if not reader.fieldnames or len(reader.fieldnames) < 2:
        raise ValueError(f"Unexpected FRED CSV header for {series_id}")
    date_column = reader.fieldnames[0]
    value_column = reader.fieldnames[1]

    for row in reader:
        value = safe_float(row.get(value_column))
        if value is None:
            continue
        try:
            observed_at = date.fromisoformat(str(row.get(date_column)))
        except ValueError:
            continue
        rows.append((observed_at, value))

    if not rows:
        raise ValueError(f"No numeric FRED observations for {series_id}")

    rows.sort(key=lambda item: item[0])
    latest_date, latest_value = rows[-1]
    prior = nearest_prior_value(rows, latest_date.replace(year=latest_date.year - 1))
    yoy = latest_value / prior[1] - 1.0 if prior and prior[1] else None
    return {
        "latest_date": latest_date.isoformat(),
        "latest_value": latest_value,
        "prior_year_date": prior[0].isoformat() if prior else None,
        "prior_year_value": prior[1] if prior else None,
        "yoy": yoy,
        "row_count": len(rows),
    }


def fetch_fred_source(source: dict[str, Any], run_dir: Path, timeout_seconds: int) -> dict[str, Any]:
    series_id = source["series_id"]
    url = source.get("url") or FRED_BASE_URL.format(series_id=series_id)
    raw_path = run_dir / "sources" / "raw" / f"{source['source_id']}.csv"
    parsed_path = run_dir / "sources" / "parsed" / f"{source['source_id']}.json"
    started = time.time()
    try:
        text, headers = fetch_text(url, timeout_seconds, {})
        write_text(raw_path, text)
        parsed = parse_fred_csv(text, series_id)
        write_json(parsed_path, parsed)
        return {
            "source_id": source["source_id"],
            "source_name": source["name"],
            "source_type": "fred_csv",
            "signal": source.get("signal"),
            "status": "ok",
            "url": url,
            "fetched_at": utc_now().isoformat(),
            "elapsed_seconds": round(time.time() - started, 3),
            "raw_path": str(raw_path),
            "parsed_path": str(parsed_path),
            "latest_date": parsed["latest_date"],
            "latest_value": parsed["latest_value"],
            "yoy": parsed["yoy"],
            "row_count": parsed["row_count"],
            "etag": headers.get("etag"),
            "error": None,
        }
    except Exception as error:
        return failed_observation(source, "fred_csv", url, started, error)


def normalized_sec_facts(units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in units:
        if item.get("form") not in {"10-K", "10-K/A", "10-Q", "10-Q/A", "20-F", "20-F/A"}:
            continue
        start = parse_iso_date(item.get("start"))
        end = parse_iso_date(item.get("end"))
        value = safe_float(item.get("val"))
        if start is None or end is None or value is None:
            continue
        rows.append(
            {
                **item,
                "start_date": start,
                "end_date": end,
                "duration_days": (end - start).days + 1,
                "val": abs(value),
            }
        )

    by_period: dict[tuple[date, date, str, str], dict[str, Any]] = {}
    for item in rows:
        key = (item["start_date"], item["end_date"], str(item.get("fp")), str(item.get("form")))
        if key not in by_period or str(item.get("filed", "")) > str(by_period[key].get("filed", "")):
            by_period[key] = item
    return sorted(by_period.values(), key=lambda item: (item["end_date"], item["start_date"], str(item.get("filed", ""))))


def annual_usd_facts(units: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_year: dict[int, dict[str, Any]] = {}
    for item in normalized_sec_facts(units):
        if item.get("form") not in {"10-K", "10-K/A", "20-F", "20-F/A"}:
            continue
        if item.get("fp") != "FY" or not 330 <= int(item["duration_days"]) <= 390:
            continue
        fy = item.get("fy")
        if fy is None:
            continue
        fy_int = int(fy)
        if fy_int not in by_year or str(item.get("filed", "")) > str(by_year[fy_int].get("filed", "")):
            by_year[fy_int] = item
    return [by_year[year] for year in sorted(by_year)]


def find_nearest_period(
    facts: list[dict[str, Any]],
    target_end: date,
    target_duration: int | None = None,
    *,
    max_end_days: int = 10,
    max_duration_days: int = 20,
) -> dict[str, Any] | None:
    candidates = []
    for item in facts:
        if days_between(item["end_date"], target_end) > max_end_days:
            continue
        if target_duration is not None and abs(int(item["duration_days"]) - target_duration) > max_duration_days:
            continue
        candidates.append(item)
    if not candidates:
        return None
    return max(candidates, key=lambda item: (str(item.get("filed", "")), item["duration_days"]))


def period_candidate(
    concept: str,
    item: dict[str, Any],
    period_type: str,
    value: float,
    source_items: list[dict[str, Any]],
    method: str,
) -> dict[str, Any]:
    return {
        "concept": concept,
        "period_type": period_type,
        "method": method,
        "start": item["start_date"],
        "end": item["end_date"],
        "fy": item.get("fy"),
        "fp": item.get("fp"),
        "form": item.get("form"),
        "filed": item.get("filed"),
        "duration_days": item["duration_days"],
        "value": value,
        "source_items": [
            {
                "form": source.get("form"),
                "fy": source.get("fy"),
                "fp": source.get("fp"),
                "start": source["start_date"].isoformat(),
                "end": source["end_date"].isoformat(),
                "filed": source.get("filed"),
                "value": source.get("val"),
            }
            for source in source_items
        ],
    }


def sec_capex_candidates(concept: str, facts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    annuals = [
        item
        for item in facts
        if item.get("form") in {"10-K", "10-K/A", "20-F", "20-F/A"}
        and item.get("fp") == "FY"
        and 330 <= int(item["duration_days"]) <= 390
    ]
    annuals.sort(key=lambda item: (item["end_date"], str(item.get("filed", ""))))
    for item in annuals:
        candidates.append(period_candidate(concept, item, "annual_fy", item["val"], [item], "reported_fy"))

    quarterlies = [item for item in facts if item.get("form") in {"10-Q", "10-Q/A"}]
    for item in quarterlies:
        duration = int(item["duration_days"])
        if 330 <= duration <= 390:
            candidates.append(period_candidate(concept, item, "ttm_direct", item["val"], [item], "reported_trailing_12m"))

    for ytd in quarterlies:
        ytd_duration = int(ytd["duration_days"])
        if not 70 <= ytd_duration <= 310:
            continue
        annual_base = max(
            [item for item in annuals if item["end_date"] < ytd["end_date"]],
            key=lambda item: (item["end_date"], str(item.get("filed", ""))),
            default=None,
        )
        if annual_base is None:
            continue
        prior_ytd = find_nearest_period(
            quarterlies,
            shift_year(ytd["end_date"], -1),
            ytd_duration,
            max_end_days=20,
            max_duration_days=35,
        )
        if prior_ytd is None:
            continue
        value = ytd["val"] + annual_base["val"] - prior_ytd["val"]
        candidates.append(
            period_candidate(
                concept,
                ytd,
                "ttm_computed",
                value,
                [ytd, annual_base, prior_ytd],
                "latest_ytd_plus_prior_fy_minus_prior_year_ytd",
            )
        )
    return candidates


def choose_prior_candidate(latest: dict[str, Any], candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    target = shift_year(latest["end"], -1)
    same_type = [item for item in candidates if item["concept"] == latest["concept"] and item["period_type"] == latest["period_type"] and item["end"] < latest["end"]]
    if not same_type and latest["period_type"].startswith("ttm"):
        same_type = [
            item
            for item in candidates
            if item["concept"] == latest["concept"] and item["period_type"].startswith("ttm") and item["end"] < latest["end"]
        ]
    if not same_type:
        same_type = [item for item in candidates if item["concept"] == latest["concept"] and item["end"] < latest["end"]]
    if not same_type:
        return None
    return min(same_type, key=lambda item: days_between(item["end"], target))


def parse_sec_capex(payload: dict[str, Any]) -> dict[str, Any]:
    facts = payload.get("facts", {}).get("us-gaap", {})
    candidates: list[dict[str, Any]] = []
    fact_counts: dict[str, int] = {}
    for concept in SEC_CAPEX_CONCEPTS:
        entries = normalized_sec_facts(facts.get(concept, {}).get("units", {}).get("USD", []))
        fact_counts[concept] = len(entries)
        candidates.extend(sec_capex_candidates(concept, entries))

    if not candidates:
        raise ValueError("No capex concept found in SEC company facts")

    priority = {"ttm_direct": 4, "ttm_computed": 3, "annual_fy": 2}
    latest = max(
        candidates,
        key=lambda item: (
            item["end"],
            priority.get(str(item["period_type"]), 0),
            str(item.get("filed") or ""),
            fact_counts.get(str(item["concept"]), 0),
        ),
    )
    prior = choose_prior_candidate(latest, candidates)
    latest_value = latest["value"]
    prior_value = prior["value"] if prior else None
    yoy = latest_value / prior_value - 1.0 if prior_value else None
    return {
        "entity_name": payload.get("entityName"),
        "concept": latest["concept"],
        "period_type": latest["period_type"],
        "method": latest["method"],
        "latest_start_date": latest["start"].isoformat(),
        "latest_date": latest["end"].isoformat(),
        "latest_fiscal_year": latest.get("fy"),
        "latest_fiscal_period": latest.get("fp"),
        "latest_form": latest.get("form"),
        "latest_filed": latest.get("filed"),
        "latest_value": latest_value,
        "prior_start_date": prior["start"].isoformat() if prior else None,
        "prior_date": prior["end"].isoformat() if prior else None,
        "prior_fiscal_year": prior.get("fy") if prior else None,
        "prior_fiscal_period": prior.get("fp") if prior else None,
        "prior_period_type": prior.get("period_type") if prior else None,
        "prior_value": prior_value,
        "yoy": yoy,
        "yoy_basis": "same_period_type_nearest_year_ago" if prior else None,
        "row_count": fact_counts.get(str(latest["concept"]), 0),
        "candidate_count": len(candidates),
        "source_items": latest["source_items"],
    }


def fetch_sec_source(source: dict[str, Any], run_dir: Path, timeout_seconds: int, user_agent: str) -> dict[str, Any]:
    cik = str(source["cik"]).zfill(10)
    url = SEC_COMPANY_FACTS_URL.format(cik=cik)
    raw_path = run_dir / "sources" / "raw" / f"{source['source_id']}.json"
    parsed_path = run_dir / "sources" / "parsed" / f"{source['source_id']}.json"
    started = time.time()
    try:
        text, headers = fetch_text(url, timeout_seconds, {"User-Agent": user_agent})
        write_text(raw_path, text)
        parsed = parse_sec_capex(json.loads(text))
        parsed["company"] = source.get("company")
        write_json(parsed_path, parsed)
        return {
            "source_id": source["source_id"],
            "source_name": source["name"],
            "source_type": "sec_companyfacts",
            "signal": "demand_capex_yoy",
            "status": "ok",
            "url": url,
            "fetched_at": utc_now().isoformat(),
            "elapsed_seconds": round(time.time() - started, 3),
            "raw_path": str(raw_path),
            "parsed_path": str(parsed_path),
            "latest_date": str(parsed.get("latest_date") or parsed.get("latest_fiscal_year")),
            "latest_value": parsed["latest_value"],
            "yoy": parsed["yoy"],
            "row_count": parsed["row_count"],
            "concept": parsed["concept"],
            "period_type": parsed.get("period_type"),
            "method": parsed.get("method"),
            "latest_start_date": parsed.get("latest_start_date"),
            "latest_fiscal_period": parsed.get("latest_fiscal_period"),
            "latest_form": parsed.get("latest_form"),
            "latest_filed": parsed.get("latest_filed"),
            "prior_date": parsed.get("prior_date"),
            "prior_value": parsed.get("prior_value"),
            "yoy_basis": parsed.get("yoy_basis"),
            "etag": headers.get("etag"),
            "error": None,
        }
    except Exception as error:
        return failed_observation(source, "sec_companyfacts", url, started, error)


def failed_observation(
    source: dict[str, Any],
    source_type: str,
    url: str,
    started: float,
    error: Exception,
) -> dict[str, Any]:
    return {
        "source_id": source["source_id"],
        "source_name": source.get("name") or source["source_id"],
        "source_type": source_type,
        "signal": source.get("signal"),
        "status": "error",
        "url": url,
        "fetched_at": utc_now().isoformat(),
        "elapsed_seconds": round(time.time() - started, 3),
        "raw_path": None,
        "parsed_path": None,
        "latest_date": None,
        "latest_value": None,
        "yoy": None,
        "row_count": 0,
        "etag": None,
        "error": f"{type(error).__name__}: {error}",
    }


def collect_sources(config: dict[str, Any], env: dict[str, str], run_dir: Path) -> list[dict[str, Any]]:
    timeout_seconds = int(env.get("PIPELINE_TIMEOUT_SECONDS", "12"))
    concurrency = max(1, int(env.get("PIPELINE_FETCH_CONCURRENCY", "4")))
    user_agent = env.get("SEC_USER_AGENT", "aicapex-monitor/0.1 admin@example.com")
    jobs: list[tuple[int, str, Any]] = []
    order = 0
    for source in config.get("fred_series", []):
        jobs.append((order, source["source_id"], ("fred", source)))
        order += 1
    for source in config.get("sec_company_capex", []):
        jobs.append((order, source["source_id"], ("sec", source)))
        order += 1

    observations_by_order: dict[int, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=min(concurrency, max(1, len(jobs)))) as executor:
        futures = {}
        for source_order, source_id, job in jobs:
            kind, source = job
            if kind == "fred":
                future = executor.submit(fetch_fred_source, source, run_dir, timeout_seconds)
            else:
                future = executor.submit(fetch_sec_source, source, run_dir, timeout_seconds, user_agent)
            futures[future] = (source_order, source_id)

        for future in as_completed(futures):
            source_order, source_id = futures[future]
            try:
                observation = future.result()
            except Exception as error:
                observation = {
                    "source_id": source_id,
                    "source_name": source_id,
                    "source_type": "unknown",
                    "signal": None,
                    "status": "error",
                    "url": None,
                    "fetched_at": utc_now().isoformat(),
                    "elapsed_seconds": None,
                    "raw_path": None,
                    "parsed_path": None,
                    "latest_date": None,
                    "latest_value": None,
                    "yoy": None,
                    "row_count": 0,
                    "etag": None,
                    "error": f"{type(error).__name__}: {error}",
                }
            observations_by_order[source_order] = observation
            print(
                f"Fetched {observation['source_id']}: {observation['status']} "
                f"({observation.get('elapsed_seconds', '-') }s)",
                flush=True,
            )

    return [observations_by_order[index] for index in sorted(observations_by_order)]


def average(values: list[float]) -> float | None:
    values = [value for value in values if value is not None and not math.isnan(value)]
    return sum(values) / len(values) if values else None


def build_signals(observations: list[dict[str, Any]], parameters: dict[str, Any]) -> dict[str, Any]:
    fred_by_signal = {
        item.get("signal"): item
        for item in observations
        if item.get("status") == "ok" and item.get("source_type") == "fred_csv"
    }
    capex_yoy_values = [
        float(item["yoy"])
        for item in observations
        if item.get("status") == "ok"
        and item.get("source_type") == "sec_companyfacts"
        and item.get("yoy") is not None
    ]

    demand_capex_yoy = average(capex_yoy_values)
    m2_yoy = fred_by_signal.get("liquidity_m2_yoy", {}).get("yoy")
    nfci_level = fred_by_signal.get("liquidity_nfci_level", {}).get("latest_value")
    brent_yoy = fred_by_signal.get("energy_brent_yoy", {}).get("yoy")

    demand_baseline = float(parameters.get("demand_capex_yoy_baseline", 0.22))
    m2_baseline = float(parameters.get("liquidity_m2_yoy_baseline", 0.04))
    brent_baseline = float(parameters.get("energy_brent_yoy_baseline", 0.0))

    demand_yoy = demand_capex_yoy if demand_capex_yoy is not None else demand_baseline
    m2 = float(m2_yoy) if m2_yoy is not None else m2_baseline
    nfci = float(nfci_level) if nfci_level is not None else 0.0
    brent = float(brent_yoy) if brent_yoy is not None else brent_baseline

    demand_delta = clamp((demand_yoy - demand_baseline) * 0.08, -0.03, 0.04)
    liquidity_delta = clamp((m2 - m2_baseline) * 0.16 + (-nfci) * 0.01, -0.015, 0.018)
    energy_delta = clamp(-(brent - brent_baseline) * 0.025, -0.018, 0.010)
    regionalization_delta = clamp(max(brent, 0.0) * 0.008 + max(demand_yoy - demand_baseline, 0.0) * 0.010, -0.004, 0.012)
    digestion_delta = -clamp(max(demand_yoy - demand_baseline, 0.0) * 0.08, 0.0, 0.035)

    return {
        "inputs": {
            "demand_capex_yoy": demand_capex_yoy,
            "liquidity_m2_yoy": m2_yoy,
            "liquidity_nfci_level": nfci_level,
            "energy_brent_yoy": brent_yoy,
            "successful_source_count": sum(1 for item in observations if item.get("status") == "ok"),
            "failed_source_count": sum(1 for item in observations if item.get("status") != "ok"),
        },
        "parameters": parameters,
        "driver_delta_anchors": {
            "AI adoption / demand diffusion": demand_delta,
            "Liquidity / financing": liquidity_delta,
            "Energy & geopolitics": energy_delta,
            "Trade / regionalization": regionalization_delta,
            "Cycle / digestion": digestion_delta,
        },
    }


def fade(driver_name: str, year: int) -> float:
    if driver_name == "AI adoption / demand diffusion":
        return {2027: 1.0, 2028: 0.85, 2029: 0.70, 2030: 0.55, 2031: 0.30, 2032: 0.18}.get(year, 0.0)
    if driver_name == "Liquidity / financing":
        return {2027: 1.0, 2028: 0.85, 2029: 0.65, 2030: 0.45, 2031: 0.25, 2032: 0.10}.get(year, 0.0)
    if driver_name == "Energy & geopolitics":
        return {2027: 1.0, 2028: 0.85, 2029: 0.65, 2030: 0.45, 2031: 0.25}.get(year, 0.0)
    if driver_name == "Trade / regionalization":
        return {2027: 0.50, 2028: 0.65, 2029: 0.85, 2030: 1.0, 2031: 0.65, 2032: 0.35}.get(year, 0.0)
    if driver_name == "Cycle / digestion":
        return {2031: 1.0, 2032: 0.85, 2033: 0.55, 2034: 0.25}.get(year, 0.0)
    return 0.0


def adjust_workbook(
    source_workbook: Path,
    generated_workbook: Path,
    observations: list[dict[str, Any]],
    signals: dict[str, Any],
    pipeline_id: str,
    generated_at: str,
) -> list[dict[str, Any]]:
    wb = load_workbook(source_workbook)
    if "Drivers" not in wb.sheetnames:
        raise RuntimeError("Workbook does not contain a Drivers sheet")

    ws = wb["Drivers"]
    year_cols = {int(ws.cell(3, col).value): col for col in range(2, ws.max_column + 1) if ws.cell(3, col).value}
    adjustments: list[dict[str, Any]] = []
    anchors = signals["driver_delta_anchors"]

    for driver_name, row in DRIVER_ROWS.items():
        anchor_delta = float(anchors.get(driver_name, 0.0))
        for year in YEARS:
            col = year_cols.get(year)
            if not col:
                continue
            baseline = safe_float(ws.cell(row, col).value) or 0.0
            delta = anchor_delta * fade(driver_name, year)
            adjusted = clamp(baseline + delta, -0.2, 0.14)
            ws.cell(row, col).value = round(adjusted, 6)
            adjustments.append(
                {
                    "driver_name": driver_name,
                    "year_num": year,
                    "baseline_contribution": baseline,
                    "adjusted_contribution": adjusted,
                    "delta_contribution": adjusted - baseline,
                    "signal_name": driver_name,
                    "rationale": f"Anchor delta {anchor_delta:.6f} with year fade {fade(driver_name, year):.2f}",
                }
            )

    append_dynamic_sources(wb, observations)
    write_update_log(wb, pipeline_id, generated_at, source_workbook, generated_workbook, observations, signals, adjustments)

    generated_workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(generated_workbook)
    return adjustments


def append_dynamic_sources(wb: Any, observations: list[dict[str, Any]]) -> None:
    if "Sources" not in wb.sheetnames:
        return
    ws = wb["Sources"]
    row = ws.max_row + 1
    ws.cell(row, 1).value = None
    row += 1
    for index, item in enumerate(observations, start=1):
        source_id = f"D{index:03d}"
        if item.get("status") == "ok":
            yoy = item.get("yoy")
            yoy_text = f"; y/y {yoy:.1%}" if isinstance(yoy, (int, float)) else ""
            key_fact = f"{item['source_name']}: latest {item.get('latest_value')} on {item.get('latest_date')}{yoy_text}"
        else:
            key_fact = f"{item['source_name']}: fetch failed ({item.get('error')})"
        ws.cell(row, 1).value = source_id
        ws.cell(row, 2).value = "Dynamic external data"
        ws.cell(row, 3).value = key_fact[:32000]
        ws.cell(row, 4).value = item.get("url")
        ws.cell(row, 5).value = item.get("source_type")
        ws.cell(row, 6).value = f"pipeline source_id={item.get('source_id')}; status={item.get('status')}"
        row += 1


def write_update_log(
    wb: Any,
    pipeline_id: str,
    generated_at: str,
    source_workbook: Path,
    generated_workbook: Path,
    observations: list[dict[str, Any]],
    signals: dict[str, Any],
    adjustments: list[dict[str, Any]],
) -> None:
    sheet_name = "Dynamic_Update_Log"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["Pipeline ID", pipeline_id])
    ws.append(["Generated at", generated_at])
    ws.append(["Source workbook", str(source_workbook)])
    ws.append(["Generated workbook", str(generated_workbook)])
    ws.append([])
    ws.append(["Signal", "Value"])
    for key, value in signals.get("inputs", {}).items():
        ws.append([key, value])
    ws.append([])
    ws.append([
        "Source ID",
        "Name",
        "Type",
        "Status",
        "Latest date",
        "Latest value",
        "YoY",
        "Period type",
        "Method",
        "Raw path",
        "Error",
    ])
    for item in observations:
        ws.append(
            [
                item.get("source_id"),
                item.get("source_name"),
                item.get("source_type"),
                item.get("status"),
                item.get("latest_date"),
                item.get("latest_value"),
                item.get("yoy"),
                item.get("period_type"),
                item.get("method"),
                item.get("raw_path"),
                item.get("error"),
            ]
        )
    ws.append([])
    ws.append(["Driver", "Year", "Baseline", "Adjusted", "Delta", "Rationale"])
    for item in adjustments:
        if abs(float(item["delta_contribution"])) < 1e-12:
            continue
        ws.append(
            [
                item["driver_name"],
                item["year_num"],
                item["baseline_contribution"],
                item["adjusted_contribution"],
                item["delta_contribution"],
                item["rationale"],
            ]
        )

    for col in range(1, 12):
        ws.column_dimensions[chr(64 + col)].width = 22 if col < 8 else 45


def run_importer(env: dict[str, str], generated_workbook: Path, pipeline_paths: dict[str, Path], pipeline_id: str) -> int:
    python_bin = env.get("PYTHON_BIN") or sys.executable
    import_script = PROJECT_ROOT / "scripts" / "import_aicapex_workbook.py"
    proc_env = os.environ.copy()
    proc_env.update(
        {
            "EXCEL_PATH": str(generated_workbook),
            "MODEL_RUN_STAMP": pipeline_id,
            "PIPELINE_ID": pipeline_id,
            "PIPELINE_SOURCE_WORKBOOK_PATH": str(pipeline_paths["source_workbook"]),
            "PIPELINE_GENERATED_WORKBOOK_PATH": str(generated_workbook),
            "PIPELINE_SOURCE_SNAPSHOT_PATH": str(pipeline_paths["source_snapshot"]),
            "PIPELINE_ADJUSTMENT_PATH": str(pipeline_paths["adjustments"]),
            "PIPELINE_MANIFEST_PATH": str(pipeline_paths["manifest"]),
            "PIPELINE_GENERATED_AT": pipeline_paths["generated_at_text"],
        }
    )
    print(f"Running model import with generated workbook: {generated_workbook}", flush=True)
    proc = subprocess.run([python_bin, str(import_script)], cwd=PROJECT_ROOT, env=proc_env)
    return proc.returncode


def main() -> int:
    env = merged_env()
    config = load_config(env)
    now = utc_now()
    pipeline_id = utc_stamp(now)
    generated_at = now.isoformat()

    source_workbook = project_path(env.get("EXCEL_PATH", "global_ai_investment_forecast_2026_2045_v7_memory_storage_split.xlsx"))
    if not source_workbook.exists():
        raise FileNotFoundError(source_workbook)

    archive_root = project_path(env.get("MODEL_ARCHIVE_DIR", "outputs/model_runs"))
    run_dir = archive_root / pipeline_id
    generated_workbook = run_dir / f"{source_workbook.stem}_dynamic_{pipeline_id}.xlsx"
    source_snapshot_path = run_dir / "source_snapshot.json"
    adjustment_path = run_dir / "driver_adjustments.json"
    manifest_path = run_dir / "pipeline_manifest.json"

    print(f"Dynamic pipeline ID: {pipeline_id}", flush=True)
    print(f"Source workbook: {source_workbook}", flush=True)
    print(f"Archive directory: {run_dir}", flush=True)

    observations = collect_sources(config, env, run_dir)
    parameters = config.get("model_parameters", {})
    successful_sources = sum(1 for item in observations if item.get("status") == "ok")
    min_success = int(parameters.get("min_successful_sources", 2))
    if successful_sources < min_success:
        print(f"Only {successful_sources} external sources succeeded; continuing with available signals and defaults.", flush=True)

    signals = build_signals(observations, parameters)
    adjustments = adjust_workbook(source_workbook, generated_workbook, observations, signals, pipeline_id, generated_at)

    write_json(source_snapshot_path, {"pipeline_id": pipeline_id, "generated_at": generated_at, "sources": observations})
    write_json(adjustment_path, {"pipeline_id": pipeline_id, "generated_at": generated_at, "signals": signals, "adjustments": adjustments})

    manifest = {
        "pipeline_id": pipeline_id,
        "generated_at": generated_at,
        "mode": os.environ.get("AUTO_UPDATE_MODE", "pipeline"),
        "trigger": os.environ.get("AUTO_UPDATE_TRIGGER", "manual"),
        "source_workbook_path": str(source_workbook),
        "generated_workbook_path": str(generated_workbook),
        "source_snapshot_path": str(source_snapshot_path),
        "adjustment_path": str(adjustment_path),
        "manifest_path": str(manifest_path),
        "successful_source_count": successful_sources,
        "failed_source_count": len(observations) - successful_sources,
        "driver_delta_anchors": signals["driver_delta_anchors"],
    }
    write_json(manifest_path, manifest)

    print(f"Wrote source snapshot: {source_snapshot_path}", flush=True)
    print(f"Wrote driver adjustments: {adjustment_path}", flush=True)
    print(f"Wrote generated workbook: {generated_workbook}", flush=True)
    print(f"Wrote pipeline manifest: {manifest_path}", flush=True)

    pipeline_paths: dict[str, Any] = {
        "source_workbook": source_workbook,
        "source_snapshot": source_snapshot_path,
        "adjustments": adjustment_path,
        "manifest": manifest_path,
        "generated_at_text": generated_at,
    }
    return run_importer(env, generated_workbook, pipeline_paths, pipeline_id)


if __name__ == "__main__":
    raise SystemExit(main())
